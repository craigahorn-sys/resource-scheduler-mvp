from __future__ import annotations

from io import BytesIO

import pandas as pd
from sqlalchemy import text

from .db import execute, query_df


# ── Migration helpers (run once on startup) ───────────────────────────────────

def migrate_revenue_columns(engine):
    """Adds billing columns to jobs and creates job_line_items if missing.
    Each ALTER runs in its own transaction so a duplicate-column error
    cannot poison the connection for subsequent statements.
    """
    billing_cols = [
        ("company_man",      "TEXT"),
        ("invoice_number",   "TEXT"),
        ("so_ticket_number", "TEXT"),
        ("day_rate",         "NUMERIC"),
        ("accrue",           "BOOLEAN NOT NULL DEFAULT FALSE"),
        ("ees_supervisor",   "TEXT"),
        ("customer_po",      "TEXT"),
        ("county_state",     "TEXT"),
        ("well_name",        "TEXT"),
        ("well_number",      "TEXT"),
        ("ordered_by",       "TEXT"),
        ("department",       "TEXT"),
        ("job_description",  "TEXT"),
        ("billing_type",     "TEXT NOT NULL DEFAULT 'line_item'"),
    ]
    for col, col_type in billing_cols:
        try:
            with engine.begin() as conn:
                conn.execute(text(f"ALTER TABLE jobs ADD COLUMN {col} {col_type}"))
        except Exception:
            pass  # column already exists — safe to ignore

    # CREATE TABLE IF NOT EXISTS is always safe to run unconditionally
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS job_line_items (
                id BIGSERIAL PRIMARY KEY,
                job_id BIGINT NOT NULL REFERENCES jobs(id) ON DELETE CASCADE,
                line_number INTEGER,
                description TEXT NOT NULL,
                uom TEXT,
                start_date DATE,
                end_date DATE,
                invoice_qty NUMERIC,
                unit_price NUMERIC,
                line_total NUMERIC,
                notes TEXT,
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """))


# ── Job billing field helpers ─────────────────────────────────────────────────

def update_job_billing(engine, job_id: int, data: dict):
    execute(engine, """
        UPDATE jobs
        SET ordered_by      = :ordered_by,
            invoice_number  = :invoice_number,
            so_ticket_number= :so_ticket_number,
            billing_type    = :billing_type,
            accrue          = :accrue,
            ees_supervisor  = :ees_supervisor,
            customer_po     = :customer_po,
            county_state    = :county_state,
            well_name       = :well_name,
            well_number     = :well_number,
            department      = :department,
            job_description = :job_description
        WHERE id = :job_id
    """, {**data, "job_id": int(job_id)})


# ── Line item CRUD ────────────────────────────────────────────────────────────

def get_line_items_df(engine, job_id: int | None = None) -> pd.DataFrame:
    if job_id is not None:
        return query_df(engine, """
            SELECT li.*, j.job_code, j.job_name, j.customer, j.region_code
            FROM job_line_items li
            JOIN jobs j ON li.job_id = j.id
            WHERE li.job_id = :job_id
            ORDER BY li.line_number NULLS LAST, li.id
        """, {"job_id": int(job_id)})
    return query_df(engine, """
        SELECT li.*, j.job_code, j.job_name, j.customer, j.region_code
        FROM job_line_items li
        JOIN jobs j ON li.job_id = j.id
        ORDER BY li.job_id, li.line_number NULLS LAST, li.id
    """)


def save_line_items(engine, job_id: int, rows: list[dict]):
    """Replace all line items for a job with the provided rows."""
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM job_line_items WHERE job_id = :job_id"), {"job_id": int(job_id)})
        for i, row in enumerate(rows, start=1):
            conn.execute(text("""
                INSERT INTO job_line_items
                    (job_id, line_number, description, uom, start_date, end_date,
                     invoice_qty, unit_price, line_total, notes)
                VALUES
                    (:job_id, :line_number, :description, :uom, :start_date, :end_date,
                     :invoice_qty, :unit_price, :line_total, :notes)
            """), {
                "job_id":      int(job_id),
                "line_number": i,
                "description": str(row.get("description", "") or ""),
                "uom":         str(row.get("uom", "") or ""),
                "start_date":  row.get("start_date") or None,
                "end_date":    row.get("end_date") or None,
                "invoice_qty": _num(row.get("invoice_qty")),
                "unit_price":  _num(row.get("unit_price")),
                "line_total":  _num(row.get("line_total")),
                "notes":       str(row.get("notes", "") or ""),
            })


def delete_line_item(engine, line_item_id: int):
    execute(engine, "DELETE FROM job_line_items WHERE id = :id", {"id": int(line_item_id)})


# ── Revenue report query ──────────────────────────────────────────────────────

def get_revenue_jobs_df(engine, region_code: str, month: int, year: int) -> pd.DataFrame:
    """
    Returns jobs for the given region whose active window overlaps the month,
    joined with their billing fields and computed dates.
    """
    month_start = f"{year}-{month:02d}-01"
    # last day of month via date arithmetic
    month_end = f"{year}-{month:02d}-01"

    df = query_df(engine, """
        SELECT
            j.id, j.job_code, j.job_name, j.customer, j.customer_color,
            j.location, j.region_code, j.status,
            j.job_start_date, j.job_duration_days,
            j.mob_days_before_job, j.demob_days_after_job,
            j.ordered_by, j.invoice_number, j.so_ticket_number,
            j.billing_type, j.day_rate, j.accrue, j.notes
        FROM jobs j
        WHERE j.region_code = :region_code
          AND j.status NOT IN ('Cancelled', 'Bid')
          AND (j.job_start_date, j.job_start_date + j.job_duration_days * INTERVAL '1 day')
              OVERLAPS (
                  DATE_TRUNC('month', CAST(:month_start AS date)),
                  DATE_TRUNC('month', CAST(:month_start AS date)) + INTERVAL '1 month'
              )
        ORDER BY j.job_start_date, j.id
    """, {"region_code": region_code, "month_start": month_start})
    return df


# ── Excel export ──────────────────────────────────────────────────────────────

def build_revenue_excel(jobs_df: pd.DataFrame, line_items_df: pd.DataFrame,
                        lob_label: str, month_label: str) -> bytes:
    """
    Builds a revenue accrual workbook matching the January 2026 CO tab layout exactly.
    - Rows 1-49:  Customer rollup panel (cols A-F) + Check panel (cols L-M)
    - Row 50:     Blank
    - Row 51+:    Job blocks (A/R header, col headers, job info, spacer,
                  line header, lines, summary label row, summary data row, gap)
    """
    from io import BytesIO
    from datetime import date as _date, datetime as _datetime
    import calendar
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = lob_label[:31]

    # ── Styles ────────────────────────────────────────────────────────────────
    GREEN_FILL = PatternFill("solid", fgColor="3D8840")
    GRAY_FILL  = PatternFill("solid", fgColor="EFEFEF")

    HAIR = Side(style="hair");  THIN = Side(style="thin")
    MED  = Side(style="medium"); NS   = Side(style=None)

    def _bdr(b=None, t=None, l=None, r=None):
        return Border(bottom=b or NS, top=t or NS,
                      left=l or NS,  right=r or NS)

    def _f(bold=False, size=10, color="000000"):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def _a(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _c(row, col, val=None, font=None, fill=None,
           bdr=None, align=None, fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        if font:  c.font   = font
        if fill:  c.fill   = fill
        if bdr:   c.border = bdr
        if align: c.alignment = align
        if fmt:   c.number_format = fmt
        return c

    def _mg(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1,
                       end_row=r2, end_column=c2)

    MONEY    = "$#,##0.00"
    DATE_FMT = "MM/DD/YYYY"
    NUM_FMT  = "0.##"

    # ── Column widths (A-N, K=spacer) ────────────────────────────────────────
    for col, w in [("A",9.13),("B",2.38),("C",18.38),("D",43.75),
                   ("E",18.0),("F",20.0),("G",8.75),("H",9.75),
                   ("I",9.38),("J",9.0),("K",2.13),
                   ("L",14.0),("M",14.0),("N",14.0)]:
        ws.column_dimensions[col].width = w

    # ── Uniform row height helper ─────────────────────────────────────────────
    def _rh(row, h=15.75):
        ws.row_dimensions[row].height = h

    # =========================================================================
    # PASS 1 — process jobs, accumulate totals
    # =========================================================================
    grand_total = 0.0
    customer_totals: dict[str, float] = {}

    # Job blocks start at row 51
    current_row = 51

    for job_num, (_, job) in enumerate(jobs_df.iterrows(), start=1):
        job_id    = int(job["id"])
        job_items = (
            line_items_df[line_items_df["job_id"] == job_id]
            .sort_values(["line_number", "id"])
            if not line_items_df.empty else pd.DataFrame()
        )

        # ── A/R header row ────────────────────────────────────────────────
        _rh(current_row)
        _mg(current_row, 8, current_row, 10)
        _c(current_row, 8, "A/R",
           font=_f(bold=True), fill=GREEN_FILL,
           bdr=_bdr(b=HAIR), align=_a(h="center"))
        _mg(current_row, 12, current_row, 14)
        _c(current_row, 12, "Approval - Ops Manager",
           font=_f(bold=True), fill=GREEN_FILL,
           bdr=_bdr(b=HAIR), align=_a(h="center"))
        current_row += 1

        # ── Column header row (green) ─────────────────────────────────────
        _rh(current_row)
        for col, label, h in [
            (1,"Job #","center"),(3,"LOB","center"),
            (4,"Lease Name/Number","center"),(5,"Customer","center"),
            (6,"Co. Man","center"),(8,"Initials","center"),
            (9,"Date","center"),(10,"Sent","center"),
            (12,"Initials","center"),(13,"Date","center"),(14,"Accrue","center"),
        ]:
            _c(current_row, col, label,
               font=_f(bold=True), fill=GREEN_FILL,
               bdr=_bdr(b=HAIR), align=_a(h=h))
        current_row += 1

        # ── Job info row (gray) ───────────────────────────────────────────
        _rh(current_row)
        for col, val, h in [
            (1, job_num,                                  "center"),
            (3, lob_label,                                "center"),
            (4, str(job.get("job_name",  "") or ""),      "left"),
            (5, str(job.get("customer",  "") or ""),      "center"),
            (6, str(job.get("ordered_by","") or ""),      "center"),
        ]:
            _c(current_row, col, val,
               font=_f(), fill=GRAY_FILL,
               bdr=_bdr(b=MED), align=_a(h=h))
        current_row += 1

        # ── Spacer row ────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 6.75
        current_row += 1

        # ── Line item header row (green, A=gray per template) ─────────────
        _rh(current_row)
        for col, label, fill, h in [
            (1,  "Line #",      GRAY_FILL,  "center"),
            (3,  "Item Abbr.",  GREEN_FILL, "center"),
            (4,  "Description", GREEN_FILL, "center"),
            (5,  "UOM",         GREEN_FILL, "center"),
            (8,  "Start",       GREEN_FILL, "center"),
            (9,  "End",         GREEN_FILL, "center"),
            (10, "# of Days",   GREEN_FILL, "center"),
            (12, "Invoice Qty", GREEN_FILL, "center"),
            (13, "Price",       GREEN_FILL, "center"),
            (14, "Line Total",  GREEN_FILL, "center"),
        ]:
            bdr = _bdr(b=THIN) if fill == GRAY_FILL else _bdr(b=HAIR)
            _c(current_row, col, label,
               font=_f(bold=True), fill=fill,
               bdr=bdr, align=_a(h=h))
        current_row += 1

        # ── Line item rows (gray) ─────────────────────────────────────────
        job_total = 0.0
        if job_items.empty:
            _rh(current_row)
            _mg(current_row, 1, current_row, 14)
            _c(current_row, 1, "(No line items entered yet)",
               font=_f(color="999999"), fill=GRAY_FILL,
               align=_a(h="center"))
            current_row += 1
        else:
            for li_num, li in enumerate(job_items.itertuples(), start=1):
                _rh(current_row)
                days = None
                sd = getattr(li, "start_date", None)
                ed = getattr(li, "end_date",   None)
                if pd.notna(sd) and pd.notna(ed):
                    try:
                        days = (pd.to_datetime(ed) - pd.to_datetime(sd)).days + 1
                    except Exception:
                        pass

                lt  = _num(li.line_total)
                qty = _num(li.invoice_qty)
                prc = _num(li.unit_price)
                if lt is None and qty is not None and prc is not None:
                    lt = qty * prc
                if lt is not None:
                    job_total += lt

                for col, val, h, fmt in [
                    (1,  li_num,  "center", None),
                    (4,  str(li.description or ""), "left",   None),
                    (5,  str(li.uom or ""),         "center", None),
                    (8,  pd.to_datetime(sd).date() if pd.notna(sd) else None,
                         "center", DATE_FMT),
                    (9,  pd.to_datetime(ed).date() if pd.notna(ed) else None,
                         "center", DATE_FMT),
                    (10, days,    "center", None),
                    (12, qty,     "center", NUM_FMT),
                    (13, prc,     "center", MONEY),
                    (14, lt,      "right",  MONEY),
                ]:
                    _c(current_row, col, val,
                       font=_f(), fill=GRAY_FILL,
                       bdr=_bdr(b=THIN), align=_a(h=h), fmt=fmt)
                current_row += 1

        # ── Summary label row (green) ─────────────────────────────────────
        _rh(current_row)
        for col, label in [
            (3,"Customer"),(4,"Lease County"),(5,"Invoice #"),
            (6,"Ticket #"),(12,"Day Rate"),(14,"Job Total"),
        ]:
            _c(current_row, col, label,
               font=_f(bold=True), fill=GREEN_FILL,
               bdr=_bdr(b=HAIR), align=_a(h="center"))
        current_row += 1

        # ── Summary data row (gray) ───────────────────────────────────────
        _rh(current_row)
        for col, val, h, fmt in [
            (3,  str(job.get("customer","")       or ""), "left",   None),
            (4,  str(job.get("county_state","")   or ""), "left",   None),
            (5,  str(job.get("invoice_number","") or ""), "center", None),
            (6,  str(job.get("so_ticket_number","")or ""),"center", None),
            (12, _num(job.get("day_rate")),               "right",  MONEY),
            (14, job_total,                               "right",  MONEY),
        ]:
            _c(current_row, col, val,
               font=_f(bold=True), fill=GRAY_FILL,
               bdr=_bdr(b=MED), align=_a(h=h), fmt=fmt)
        current_row += 2   # blank gap between jobs

        grand_total += job_total
        cust = str(job.get("customer","") or "Unassigned")
        customer_totals[cust] = customer_totals.get(cust, 0.0) + job_total

    # =========================================================================
    # PASS 2 — fill customer rollup panel (rows 2-49) and check panel
    # =========================================================================

    # Row 2: panel headers
    _rh(2)
    for col, label, h in [(1,"LOB","right"),(3,"Projection","center"),
                           (5,"Customer","left"),(6,"Total","center")]:
        _c(2, col, label,
           font=_f(bold=True), fill=GREEN_FILL,
           bdr=_bdr(b=HAIR), align=_a(h=h))

    # Row 3: LOB row + check panel start
    _rh(3)
    _c(3, 1, lob_label,
       font=_f(bold=True), fill=GRAY_FILL,
       bdr=_bdr(b=HAIR), align=_a(h="right"))
    _c(3, 3, grand_total,
       font=_f(bold=True), fill=GRAY_FILL,
       bdr=_bdr(b=HAIR), align=_a(h="center"), fmt=MONEY)
    # Check panel labels
    _c(3, 12, "Check", font=_f(), align=_a(h="left"))

    # Check panel: Revenue / Customer / Delta
    _c(4, 12, "Revenue",   font=_f(), align=_a(h="left"))
    _c(4, 13, grand_total, font=_f(), align=_a(h="right"), fmt=MONEY)
    _c(5, 12, "Customer",  font=_f(), align=_a(h="left"))
    _c(5, 13, grand_total, font=_f(), align=_a(h="right"), fmt=MONEY)
    _c(6, 12, "Delta",     font=_f(), align=_a(h="left"))
    _c(6, 13, 0.0,         font=_f(), align=_a(h="right"), fmt=MONEY)

    # Customer list rows 4-49
    panel_row = 4
    for cust, total in sorted(customer_totals.items()):
        if panel_row > 49:
            break
        _rh(panel_row)
        _c(panel_row, 5, cust,
           font=_f(), fill=GRAY_FILL,
           bdr=_bdr(b=HAIR), align=_a(h="left"))
        _c(panel_row, 6, total,
           font=_f(), fill=GRAY_FILL,
           bdr=_bdr(b=HAIR), align=_a(h="center"), fmt=MONEY)
        panel_row += 1

    # Grand total row in panel
    if panel_row <= 49:
        _rh(panel_row)
        _c(panel_row, 5, "TOTAL",
           font=_f(bold=True), fill=GREEN_FILL,
           bdr=_bdr(b=HAIR), align=_a(h="left"))
        _c(panel_row, 6, grand_total,
           font=_f(bold=True), fill=GREEN_FILL,
           bdr=_bdr(b=HAIR), align=_a(h="center"), fmt=MONEY)

    # Total line below panel (row 49 of original)
    _c(49, 6, grand_total,
       font=_f(bold=True), fill=GRAY_FILL,
       bdr=_bdr(b=MED), align=_a(h="right"), fmt=MONEY)

    ws.freeze_panes = "A51"

    # =========================================================================
    # PASS 3 — Summary sheet
    # =========================================================================
    from datetime import datetime as _dt
    today = _date.today()
    try:
        month_dt = _dt.strptime(month_label, "%B %Y")
        report_month, report_year = month_dt.month, month_dt.year
    except Exception:
        report_month, report_year = today.month, today.year

    days_in_month = calendar.monthrange(report_year, report_month)[1]
    days_elapsed  = (today.day
                     if report_month == today.month and report_year == today.year
                     else days_in_month)

    ss = wb.create_sheet("Summary")
    for col, w in [("A",28),("B",18),("C",14),
                   ("D",18),("E",18),("F",16)]:
        ss.column_dimensions[col].width = w

    # Title
    ss.merge_cells("A1:F1")
    c = ss["A1"]
    c.value = f"{lob_label}  —  Monthly Summary  |  {month_label}"
    c.font  = _f(bold=True, size=12, color="FFFFFF")
    c.fill  = GREEN_FILL
    c.alignment = _a(h="center")
    ss.row_dimensions[1].height = 22

    # Metadata
    ss.merge_cells("A2:F2")
    c = ss["A2"]
    c.value = (f"Days Elapsed: {days_elapsed}  of  {days_in_month}"
               f"  |  Generated: {today.strftime('%m/%d/%Y')}")
    c.font  = _f(color="444444")
    c.alignment = _a(h="center")
    ss.row_dimensions[2].height = 16

    # Column headers
    for ci, label in enumerate(
        ["Customer","Actual Total","Days Elapsed",
         "Daily Run Rate","Projected Total","Variance"], start=1):
        c = ss.cell(row=3, column=ci, value=label)
        c.font   = _f(bold=True, color="FFFFFF")
        c.fill   = GREEN_FILL
        c.border = _bdr(b=THIN)
        c.alignment = _a(h="center")
    ss.row_dimensions[3].height = 22

    # Customer rows
    for ri, (cust, actual) in enumerate(sorted(customer_totals.items()), start=4):
        run_rate  = actual / days_elapsed if days_elapsed > 0 else 0.0
        projected = run_rate * days_in_month
        variance  = projected - actual
        ss.row_dimensions[ri].height = 15.75
        for ci, val, fmt, h, bold, color in [
            (1, cust,       None,  "left",   False, "000000"),
            (2, actual,     MONEY, "right",  False, "000000"),
            (3, days_elapsed,"0",  "center", False, "000000"),
            (4, run_rate,   MONEY, "right",  False, "000000"),
            (5, projected,  MONEY, "right",  False, "000000"),
            (6, variance,   MONEY, "right",  True,
             "375623" if variance >= 0 else "9C0006"),
        ]:
            c = ss.cell(row=ri, column=ci, value=val)
            c.font   = _f(bold=bold, color=color)
            c.fill   = GRAY_FILL
            c.border = _bdr(b=HAIR)
            c.alignment = _a(h=h)
            if fmt: c.number_format = fmt

    # Grand total
    tr = 4 + len(customer_totals)
    ss.row_dimensions[tr].height = 18
    gr = grand_total / days_elapsed if days_elapsed > 0 else 0.0
    gp = gr * days_in_month
    gv = gp - grand_total
    for ci, val, fmt, h in [
        (1, "TOTAL",      None,  "left"),
        (2, grand_total,  MONEY, "right"),
        (3, days_elapsed, "0",   "center"),
        (4, gr,           MONEY, "right"),
        (5, gp,           MONEY, "right"),
        (6, gv,           MONEY, "right"),
    ]:
        c = ss.cell(row=tr, column=ci, value=val)
        c.font   = _f(bold=True, color="FFFFFF")
        c.fill   = GREEN_FILL
        c.border = _bdr(b=MED)
        c.alignment = _a(h=h)
        if fmt: c.number_format = fmt

    ss.freeze_panes = "A4"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

def _num(val):
    """Convert to float or return None."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        return float(val)
    except Exception:
        return None


def _fmt(val) -> str:
    n = _num(val)
    return f"{n:,.2f}" if n is not None else "—"


# ── Field Ticket Excel export ─────────────────────────────────────────────────

def build_ticket_excel(job: dict, line_items_df) -> bytes:
    """
    Loads 2025_Ticket_Sample.xlsx template, writes job data into specific cells.
    Only writes B (qty) and G (unit price) for line items.
    H column =G*B formulas and H31 SUM are baked into the template — never touched.
    """
    from io import BytesIO
    from pathlib import Path
    from openpyxl import load_workbook

    DOLLAR_FMT = "$#,##0.00"

    template_path = Path(__file__).resolve().parent.parent / "2025_Ticket_Sample.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(
            f"Field ticket template not found at {template_path}. "
            "Please add '2025_Ticket_Sample.xlsx' to the repo root."
        )

    wb = load_workbook(template_path)
    ws = wb.active

    # Clear data cells only — preserves all formatting, borders, and H formulas
    cells_to_clear = (
        ["H1", "B11"] +
        [f"C{r}" for r in [6]] +           # Customer value (merged C6:D6)
        [f"F{r}" for r in [6]] +           # Date value (merged F6:H6)
        [f"B{r}" for r in [7, 8, 9, 10]] + # Left value cells (merged B:D)
        [f"F{r}" for r in [7, 8, 9, 10]] + # Right value cells (merged F:H)
        [f"B{r}" for r in range(13, 31)] +  # Line item qty
        [f"C{r}" for r in range(13, 31)] +  # Line item uom
        [f"D{r}" for r in range(13, 31)] +  # Line item description
        [f"G{r}" for r in range(13, 31)]    # Line item unit price
        # H column intentionally omitted — template formulas stay intact
    )
    for coord in cells_to_clear:
        try:
            ws[coord].value = None
        except Exception:
            pass

    import math as _math
    import pandas as _pd

    def _cell(val):
        """Safely convert a value to string, treating NaN/None as empty."""
        if val is None:
            return ""
        try:
            if isinstance(val, float) and (_math.isnan(val) or _math.isinf(val)):
                return ""
        except Exception:
            pass
        s = str(val).strip()
        return "" if s in ("nan", "None", "NaT", "NaN") else s

    def _date(val):
        """Format a date value as MM/DD/YYYY."""
        if val is None:
            return ""
        try:
            if isinstance(val, float) and _math.isnan(val):
                return ""
            ts = _pd.to_datetime(val, errors="coerce")
            if _pd.isna(ts):
                return ""
            return ts.strftime("%m/%d/%Y")
        except Exception:
            return str(val)

    # Row 1: SO / Ticket number
    ws["H1"].value = _cell(job.get("so_ticket_number"))

    # Row 6: Customer Name | Date of Service
    ws["C6"].value = _cell(job.get("customer"))
    ws["F6"].value = _date(job.get("job_start_date"))

    # Row 7: Ordered By | Job Name
    ws["B7"].value = _cell(job.get("ordered_by"))
    ws["F7"].value = _cell(job.get("job_name"))

    # Row 8: Customer PO | County, State
    ws["B8"].value = _cell(job.get("customer_po"))
    ws["F8"].value = _cell(job.get("county_state"))

    # Row 9: Well Name | EES Supervisor
    ws["B9"].value = _cell(job.get("well_name"))
    ws["F9"].value = _cell(job.get("ees_supervisor"))

    # Row 10: Well Number | Department
    ws["B10"].value = _cell(job.get("well_number"))
    ws["F10"].value = _cell(job.get("department"))

    # Row 11: Job Description
    ws["B11"].value = _cell(job.get("job_description"))

    # Line items: write B (qty) and G (unit price) only
    # H column keeps =G*B formulas — do NOT write to H
    items = line_items_df.to_dict("records") if not line_items_df.empty else []
    for i in range(18):
        row = 13 + i
        if i < len(items):
            li = items[i]
            qty   = _num(li.get("invoice_qty"))
            price = _num(li.get("unit_price"))
            ws[f"B{row}"].value = qty
            ws[f"C{row}"].value = str(li.get("uom", "") or "")
            ws[f"D{row}"].value = str(li.get("description", "") or "")
            ws[f"G{row}"].value = price

    # Dollar formatting on G13:G30 only
    for r in range(13, 31):
        ws[f"G{r}"].number_format = DOLLAR_FMT

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
